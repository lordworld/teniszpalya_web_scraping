import string

import openpyxl
import os
from openpyxl.styles import Font, Color
from openpyxl.styles import PatternFill

class ExcelManagerClass:

    def __init__(self, name, path = os.getcwd()):
        self.name = name
        self.path = path
        self.workbook = openpyxl.Workbook()

    def create_sheet(self, name):
        self.sheet = self.workbook.create_sheet(name)

    def select_sheet(self, name):
        self.sheet = self.workbook[name]

    def get_sheet_names(self):
        return self.workbook.get_sheet_names()

    def load_existing_excel(self, path):
        self.workbook = openpyxl.load_workbook(path)

    def __set_path(self, path):
        self.path = path

    def change_path(self, path):
        self.set_path(path)
        os.ch.dir(path)

    def close_excel(self):
        self.workbook.close()

    def get_sheet_by_name(self, name):
        return self.workbook.get_sheet_by_name(name)

    def color_cells(self, first_cells, last_cell, patternType = 'solid', fgColor = 'D9D9D9', bgColor = Color()):
        for column_num in range(ord(first_cells[0]),ord(last_cell[0]) + 1):
            for row in range(int(first_cells[1:]),int(last_cell[1:]) + 1):
                column = chr(column_num)
                cell = column + str(row)
                fill_pattern = PatternFill(patternType = patternType, fgColor= fgColor, bgColor=bgColor)
                self.sheet[cell].fill = fill_pattern

    def remove_sheet_by_name(self, name):
        sheet_to_remove = self.get_sheet_by_name(name)
        print(sheet_to_remove)
        self.workbook.remove_sheet(sheet_to_remove)

    def save_excel(self, path = False):
        if not path:
            path = self.path
        else:
            self.change_path(path)

        name = self.name
        # print(name)
        self.workbook.save(name)

    def set_excel_header_names(self, names):
        i = 0
        for name in names:
            letter = chr(ord('A') + i)
            cell = letter + '1'
            i += 1
            self.sheet[cell] = name
        pass

    def set_excel_row_size(self, sizes):
        i = 0
        for size in sizes:
            letter = chr(ord('A') + i)
            i += 1
            self.sheet.column_dimensions[letter].width = size

    def fill_excel_cell(self, cell, data):
        self.sheet[cell] = data